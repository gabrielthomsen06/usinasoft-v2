import io
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_db, get_current_admin_user
from app.models.usuario import Usuario
from app.schemas.dashboard_financeiro import DashboardFinanceiro
from app.services.dashboard_financeiro_service import get_dashboard_data, get_export_data

router = APIRouter(prefix="/dashboard-financeiro", tags=["dashboard-financeiro"])

MONTHS_PT = [
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

CATEGORIA_LABELS = {
    "material": "Material",
    "servicos": "Serviços",
    "fixas": "Fixas",
    "impostos": "Impostos",
    "carro": "Carro",
    "gasolina": "Gasolina",
    "salario": "Salário",
    "aluguel": "Aluguel",
    "patrimonio": "Patrimônio",
    "outros": "Outros",
}


@router.get("/", response_model=DashboardFinanceiro)
async def get_dashboard(
    mes: Optional[int] = Query(None, ge=1, le=12),
    ano: Optional[int] = Query(None, ge=2020, le=2100),
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(get_current_admin_user),
) -> dict:
    return await get_dashboard_data(db, mes=mes, ano=ano)


@router.get("/exportar")
async def exportar_excel(
    mes: Optional[int] = Query(None, ge=1, le=12),
    ano: Optional[int] = Query(None, ge=2020, le=2100),
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(get_current_admin_user),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = await get_export_data(db, mes=mes, ano=ano)
    periodo = f"{MONTHS_PT[data['mes']]} {data['ano']}"

    wb = openpyxl.Workbook()

    # Estilos
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A2340", end_color="1A2340", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    title_font = Font(name="Calibri", size=14, bold=True, color="1A2340")
    subtitle_font = Font(name="Calibri", size=10, color="888888")
    currency_fmt = '#,##0.00'
    date_fmt = 'DD/MM/YYYY'
    thin_border = Border(
        bottom=Side(style="thin", color="E0E0E0"),
    )
    gold_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def auto_width(ws, cols):
        for col in range(1, cols + 1):
            max_len = 0
            for row in ws.iter_rows(min_col=col, max_col=col, min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    # ===== ABA 1: Lançamentos =====
    ws1 = wb.active
    ws1.title = "Lançamentos"
    ws1.cell(row=1, column=1, value=f"Lançamentos — {periodo}").font = title_font
    ws1.merge_cells("A1:F1")

    headers = ["Data", "Tipo", "Descrição", "Valor (R$)"]
    row = 3
    for i, h in enumerate(headers, 1):
        ws1.cell(row=row, column=i, value=h)
    style_header(ws1, row, len(headers))

    total_rec = 0
    total_desp = 0
    for lanc in data["lancamentos"]:
        row += 1
        ws1.cell(row=row, column=1, value=lanc.data).number_format = date_fmt
        tipo_label = "Receita" if lanc.tipo == "receita" else "Despesa"
        ws1.cell(row=row, column=2, value=tipo_label)
        ws1.cell(row=row, column=3, value=lanc.descricao)
        val = float(lanc.valor)
        ws1.cell(row=row, column=4, value=val).number_format = currency_fmt
        for col in range(1, 5):
            ws1.cell(row=row, column=col).border = thin_border
        if lanc.tipo == "receita":
            total_rec += val
        else:
            total_desp += val

    # Totais
    row += 2
    ws1.cell(row=row, column=3, value="Total Receitas:").font = Font(bold=True)
    ws1.cell(row=row, column=4, value=total_rec).number_format = currency_fmt
    ws1.cell(row=row, column=4).font = Font(bold=True, color="2E7D32")
    row += 1
    ws1.cell(row=row, column=3, value="Total Despesas:").font = Font(bold=True)
    ws1.cell(row=row, column=4, value=total_desp).number_format = currency_fmt
    ws1.cell(row=row, column=4).font = Font(bold=True, color="C62828")
    row += 1
    ws1.cell(row=row, column=3, value="Saldo:").font = Font(bold=True, size=12)
    ws1.cell(row=row, column=4, value=total_rec - total_desp).number_format = currency_fmt
    ws1.cell(row=row, column=4).font = Font(bold=True, size=12, color="1A2340")
    ws1.cell(row=row, column=4).fill = gold_fill

    auto_width(ws1, 4)

    # ===== ABA 2: Contas a Pagar =====
    ws2 = wb.create_sheet("Contas a Pagar")
    ws2.cell(row=1, column=1, value=f"Contas a Pagar — {periodo}").font = title_font
    ws2.merge_cells("A1:G1")

    headers = ["Descrição", "Fornecedor", "Categoria", "Valor (R$)", "Vencimento", "Status", "Parcela"]
    row = 3
    for i, h in enumerate(headers, 1):
        ws2.cell(row=row, column=i, value=h)
    style_header(ws2, row, len(headers))

    for cp in data["contas_pagar"]:
        row += 1
        ws2.cell(row=row, column=1, value=cp.descricao)
        ws2.cell(row=row, column=2, value=cp.fornecedor.nome if cp.fornecedor else "—")
        ws2.cell(row=row, column=3, value=CATEGORIA_LABELS.get(cp.categoria, cp.categoria))
        ws2.cell(row=row, column=4, value=float(cp.valor)).number_format = currency_fmt
        ws2.cell(row=row, column=5, value=cp.data_vencimento).number_format = date_fmt
        ws2.cell(row=row, column=6, value=cp.status.capitalize())
        parcela = f"{cp.parcela_atual}/{cp.total_parcelas}" if cp.total_parcelas > 1 else "—"
        ws2.cell(row=row, column=7, value=parcela)
        for col in range(1, 8):
            ws2.cell(row=row, column=col).border = thin_border

    auto_width(ws2, 7)

    # ===== ABA 3: Contas a Receber =====
    ws3 = wb.create_sheet("Contas a Receber")
    ws3.cell(row=1, column=1, value=f"Contas a Receber — {periodo}").font = title_font
    ws3.merge_cells("A1:F1")

    headers = ["Descrição", "Cliente", "Valor (R$)", "Vencimento", "Pagamento", "Status"]
    row = 3
    for i, h in enumerate(headers, 1):
        ws3.cell(row=row, column=i, value=h)
    style_header(ws3, row, len(headers))

    for cr in data["contas_receber"]:
        row += 1
        ws3.cell(row=row, column=1, value=cr.descricao)
        ws3.cell(row=row, column=2, value=cr.cliente.nome if cr.cliente else "—")
        ws3.cell(row=row, column=3, value=float(cr.valor)).number_format = currency_fmt
        ws3.cell(row=row, column=4, value=cr.data_vencimento).number_format = date_fmt
        ws3.cell(row=row, column=5, value=cr.data_pagamento or "—")
        ws3.cell(row=row, column=6, value=cr.status.capitalize())
        for col in range(1, 7):
            ws3.cell(row=row, column=col).border = thin_border

    auto_width(ws3, 6)

    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Financeiro_{MONTHS_PT[data['mes']]}_{data['ano']}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
